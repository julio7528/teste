import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32
import re

def convert_to_xlsx(file_path):
    """Converte o arquivo para .xlsx novamente para corrigir possíveis problemas de arquivo."""
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(file_path)
        file_path_xlsx = file_path  # Adiciona _reparado para evitar sobrescrever
        wb.SaveAs(file_path_xlsx, FileFormat=51)  # 51 = formato .xlsx
        wb.Close()
        return file_path_xlsx
    except Exception as e:
        print(f"Erro ao converter o arquivo: {e}")
        return file_path
    finally:
        excel.Quit()


def verify_excel_type(file_path):
    try:

        if file_path.endswith(".xlsx"):
            file_path = convert_to_xlsx(file_path)

        df = pd.read_excel(file_path, header=None)
        
        # Primeira célula que queremos verificar
        cell_title_position_1 = df.at[1, 0]
        matches = False

        # Verifica se cell_title_position_1 é string e contém as palavras-chave esperadas
        if isinstance(cell_title_position_1, str):
            if "LISTA DE" in cell_title_position_1:
                matches = True
                return "TYPE_A"
            elif "DADOS FORNECEDOR" in cell_title_position_1:
                matches = True
                return "TYPE_B"
            
        if matches == False:
            # Verificar se a palavra "revisão" aparece em alguma célula da coluna A
            for value in df[0]:  # Coluna A é o índice 0
                if isinstance(value, str) and "revisão" in value.lower():
                    matches = True
                    return "TYPE_REVISION"
                
        if matches == False:
            # Carregar o arquivo Excel para verificar o rodapé
            wb = load_workbook(file_path)
            sheet = wb.active

            rodape_esquerdo = str(sheet.oddFooter.left.text) if sheet.oddFooter.left.text else ""
            match = re.search(r"(Revisão\s+)(\d+)", rodape_esquerdo)
            if match:
                matches = True
                return "TYPE_REVISION"

        if matches == False:
        
            # Se nenhuma das condições acima for satisfeita, verificar outra célula
            cell_title_position_2 = df.at[0, 1]
        
            if isinstance(cell_title_position_2, str):
                if  "PLANILHA DE CÁLCULOS" in cell_title_position_2:
                    matches = True
                    return "TYPE_C"
     
        if matches == False:

            wb = load_workbook(file_path)
            # Acessar a primeira planilha
            sheet = wb.active  # ou wb['NomeDaPlanilha'] se quiser especificar uma aba

            # Tentar acessar os cabeçalhos do layout de página
            cabecalho_esquerdo = getattr(sheet, "oddHeader", None)
            cabecalho_central = getattr(sheet, "evenHeader", None)
            cabecalho_direito = getattr(sheet, "firstHeader", None)

            # Filtrar apenas os cabeçalhos que possuem conteúdo
            cabecalhos = [str(c).strip() for c in [cabecalho_esquerdo, cabecalho_central, cabecalho_direito] if c]

            if any("FICHA DE CÁLCULOS" in c for c in cabecalhos):
                return "TYPE_D"

        return "INVALID_TYPE"

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")
        return None


