import pandas as pd
import re
from openpyxl import load_workbook
from xlsxwriter import Workbook as XWorkBook
from datetime import datetime
from openpyxl.styles import Alignment
import win32com.client as win32

def buscar_descricao_para(codigo_de, df):
    codigo_de_formatado = str(codigo_de).split(" ")[0]
    try:
        resultado = df.loc[df['Codigo - DE'].astype(str) == str(codigo_de_formatado)]
        if not resultado.empty:
            return str(resultado['Descrição Item'].iloc[0])
        return ""
    except Exception as e:
        return f'Erro: {e}'
# Função para buscar na tabela DePara
def buscar_codigo_para(codigo_de, df):
    codigo_de_formatado = str(codigo_de).split(" ")[0]
    try:
        resultado = df.loc[df['Codigo - DE'].astype(str) == str(codigo_de_formatado)]
        if not resultado.empty:
            return str(resultado['Codigo -  PARA'].iloc[0])
        return 'Código não encontrado'
    except Exception as e:
        return f'Erro: {e}'
    
def edit_excel_codigo(arquivo_principal, arquivo_consulta, excel_type_verification):
    # Carregar o arquivo principal sem cabeçalhos
    df = pd.read_excel(arquivo_principal, header=None)

    # Carregar o arquivo DePara
    df_consulta = pd.read_excel(arquivo_consulta, header=2)

    # Capturar a coluna A (índice 0) a partir da linha 3 (índice 2 em pandas)
    coluna_a = df.iloc[3:,0 if excel_type_verification == "TYPE_A" else 5]

    # Encontrar a posição onde está "REVISÃO" e parar antes dessa linha
    posicao_parada = coluna_a[coluna_a.astype(str).str.contains("REVISÃO", na=False)].index.min()

    # Ajustar para capturar apenas até a linha antes da palavra "REVISÃO"
    if pd.notna(posicao_parada):  # Verificar se a posição de parada é válida
        coluna_a = coluna_a.iloc[:(posicao_parada-1)].reset_index(drop=True)
    else:
        print("O valor 'REVISÃO' não foi encontrado. Capturando toda a coluna.")
        coluna_a = coluna_a.reset_index(drop=True)

    # Remover células vazias ou inválidas
    coluna_a = coluna_a.dropna().astype(str)

    # Processar códigos com \n e buscar na tabela DePara
    def processar_multiplos_codigos(valor, df):
        codigos = str(valor).split('\n')  # Dividir os códigos
        resultados = []
        
        if excel_type_verification == "TYPE_A":
            resultados = [buscar_codigo_para(codigo.strip(), df) for codigo in codigos]
        else:
            resultados = [f"{buscar_codigo_para(codigo.strip(), df)} {buscar_descricao_para(codigo.strip(), df)}" for codigo in codigos]
        return '\n'.join(map(str, resultados))  # Garantir que todos os resultados são strings

    # Aplicar a função para processar múltiplos códigos
    coluna_b = coluna_a.apply(
        lambda valor: processar_multiplos_codigos(valor, df_consulta)
    )

    # Carregar o arquivo Excel original com openpyxl para preservar a formatação
    wb = load_workbook(arquivo_principal)
    ws = wb.active
    
    coluna_b_index = 2 if excel_type_verification == "TYPE_A" else 7

    # Atualizar somente a coluna B (índice 1 no Excel, começando na linha 3)
    for i, valor in enumerate(coluna_b, start=4):
        # Verificar se a célula correspondente na coluna A não está vazia
        if ws.cell(row=i, column=1).value:  # Apenas atualiza se a célula da coluna A não estiver vazia
            # Verificar se a célula é mesclada
            if not ws.cell(row=i, column=coluna_b_index).coordinate in ws.merged_cells:
                ws.cell(row=i, column=coluna_b_index, value=valor)  # Atualizar somente células não mescladas
            else:
                # Identificar a célula superior esquerda do merge e atualizar
                for merged_range in ws.merged_cells.ranges:
                    if ws.cell(row=i, column=2).coordinate in merged_range:
                        top_left = merged_range.min_row, merged_range.min_col
                        ws.cell(row=top_left[0], column=top_left[1], value=valor)
                        break

    # Salvar o arquivo preservando a formatação
    wb.save(arquivo_principal)

def adicionar_revisao(arquivo_principal, motivo, data=None):
    """
    Adiciona uma nova revisão ao arquivo Excel com base na última revisão encontrada,
    insere uma nova linha, mescla células conforme o padrão e alinha os textos ao centro.

    Args:
        arquivo_principal (str): Caminho para o arquivo Excel.
        motivo (str): Motivo da nova revisão.
        data (str): Data da nova revisão no formato 'dd/mm/aaaa'. Se não for fornecida, usa a data atual.
    """
    # Carregar o arquivo Excel original com openpyxl para preservar a formatação
    wb = load_workbook(arquivo_principal)
    ws = wb.active

    # Identificar a linha onde começam as revisões (procurar pela palavra "REVISÃO")
    linha_revisoes = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "REVISÃO" in str(cell.value).upper():
                linha_revisoes = cell.row
                break
        if linha_revisoes:
            break

    if not linha_revisoes:
        print("Não foi possível encontrar a seção de revisões.")
        return

    # Encontrar a última linha com revisão
    ultima_linha = linha_revisoes + 1
    ultima_revisao = 0  # Inicializar com 0 caso não haja revisões anteriores

    while True:
        valor_celula = ws.cell(row=ultima_linha, column=1).value  # Assume que a coluna 1 contém os números das revisões
        if valor_celula is None:  # Se a célula estiver vazia, é o fim das revisões
            break
        try:
            # Tentar converter para inteiro para identificar o número da última revisão
            ultima_revisao = int(valor_celula)
        except ValueError:
            pass  # Ignorar valores não numéricos
        ultima_linha += 1

    # Inserir uma nova linha
    ws.insert_rows(ultima_linha)

    # Determinar o número da nova revisão
    nova_revisao = ultima_revisao + 1

    # Usar a data atual se nenhuma for fornecida
    if not data:
        data = datetime.now().strftime('%d/%m/%Y')

    # Adicionar a nova revisão
    # Mesclar A, B, C para a nova revisão
    ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=3)
    celula_revisao = ws.cell(row=ultima_linha, column=1, value=nova_revisao)  # Coluna A (mesclada com B e C)

    # Mesclar D, E, F para o motivo
    ws.merge_cells(start_row=ultima_linha, start_column=4, end_row=ultima_linha, end_column=6)
    celula_motivo = ws.cell(row=ultima_linha, column=4, value=motivo)  # Coluna D (mesclada com E e F)

    # Colocar a data na coluna G
    celula_data = ws.cell(row=ultima_linha, column=7, value=data)  # Coluna G (não mesclada)

    # Centralizar o texto nas células mescladas e na data
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    celula_revisao.alignment = alinhamento_centro
    celula_motivo.alignment = alinhamento_centro
    celula_data.alignment = alinhamento_centro

    # Salvar o arquivo preservando a formatação
    wb.save(arquivo_principal)
    print(f"Revisão {nova_revisao} adicionada com sucesso.")


def get_excel_footer(file_path):
    try:
        # Carregar o arquivo Excel
        wb = load_workbook(file_path)
        sheet = wb.active  # Obtém a primeira aba ativa

        # Capturar os rodapés corretamente convertendo para string
        rodape_esquerdo = str(sheet.oddFooter) if sheet.oddFooter else None
        rodape_central = str(sheet.evenFooter) if sheet.evenFooter else None
        rodape_direito = str(sheet.firstFooter) if sheet.firstFooter else None

        # Filtrar apenas os rodapés que possuem conteúdo
        rodapes = [r.strip() for r in [rodape_esquerdo, rodape_central, rodape_direito] if r]

        if rodapes:

            # sheet.oddFooter.left.text = "banana0"
            

            return " | ".join(rodapes)  # Junta os rodapés encontrados
        else:
            return "Nenhum rodapé encontrado"

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")
        return None


def convert_to_xlsx(file_path):
    """Converte o arquivo para .xlsx novamente para corrigir possíveis problemas de arquivo."""
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(file_path)
        file_path_xlsx = file_path  # Adiciona _reparado para evitar sobrescrever
        wb.SaveAs(file_path_xlsx, FileFormat=51)  # 51 = formato .xlsx
        wb.Close()
        return file_path_xlsx
    except Exception as e:
        print(f"Erro ao converter o arquivo: {e}")
        return file_path
    finally:
        excel.Quit()

def update_excel_footer(file_path):
    try:
        # Converter para .xlsx novamente (mesmo que já seja .xlsx)
        #função criada para bug de carregamento de workbook do excel file
        if file_path.endswith(".xlsx"):
            file_path = convert_to_xlsx(file_path)

        # Carregar o arquivo Excel para verificar o rodapé
        wb = load_workbook(file_path)
        sheet = wb.active

        rodape_esquerdo = str(sheet.oddFooter.left.text) if sheet.oddFooter.left.text else ""
        match = re.search(r"(Revisão\s+)(\d+)", rodape_esquerdo)

        if match:
            # Atualiza a revisão no rodapé
            prefixo = match.group(1)
            numero_atual = int(match.group(2))
            novo_numero = numero_atual + 1
            novo_rodape = rodape_esquerdo.replace(f"{prefixo}{numero_atual:02}", f"{prefixo}{novo_numero:02}")
            sheet.oddFooter.left.text = novo_rodape
            wb.save(file_path)
            print(f"Rodapé atualizado: {novo_rodape}")
        else:
            print("Número de revisão não encontrado no rodapé. Verificando na coluna A...")

            # Verificar revisão na coluna A usando openpyxl
            for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1):
                for cell in row:
                    # Imprimir o valor de cada célula para ver o que está sendo processado
                    print(f"Verificando célula: {cell.value}")

                    if isinstance(cell.value, str):
                        # A expressão regular agora está procurando por "revisão" seguido de um número após uma vírgula
                        match_col_a = re.search(r"(revisão\s+)(\d+)", cell.value, re.IGNORECASE)
                        if match_col_a:
                            # Construa o novo texto preservando o resto da string, mas substituindo o número da revisão
                            novo_texto = cell.value.replace(match_col_a.group(0), f"{match_col_a.group(1)}{int(match_col_a.group(2)) + 1:02}")
                            cell.value = novo_texto
                            print(f"Revisão encontrada e alterada: {novo_texto}")

            # Salvar o arquivo atualizado na coluna A
            wb.save(file_path)
            print("Revisões atualizadas na coluna A e no rodapé.")

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")

